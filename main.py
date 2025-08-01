from fastapi import FastAPI, Request, Form, Query
from fastapi.responses import RedirectResponse, FileResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from database import SessionLocal, engine
from models import Base, Prayer
from datetime import datetime, timedelta
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment
import io
import os
import pytz  # ✅ 추가
from dotenv import load_dotenv
from fastapi import HTTPException, status
from version import get_version, get_build_info  # 버전 정보 import
import time  # 캐싱용
from typing import Optional  # 타입 힌트용

# .env 파일 로드 (개발환경에서만)
load_dotenv()

# ✅ 한국 시간대 설정
KST = pytz.timezone("Asia/Seoul")

# ✅ FastAPI 앱 인스턴스 생성
app = FastAPI()

# 📦 캐싱을 위한 간단한 메모리 저장소
view_cache = {
    "data": None,
    "timestamp": 0,
    "ttl": 30  # 30초 캐싱
}

# ✅ Static 파일 설정
app.mount("/static", StaticFiles(directory="static"), name="static")

# ✅ Jinja 템플릿 설정
templates = Jinja2Templates(directory="templates")

# ✅ DB 초기화
Base.metadata.create_all(bind=engine)


# ✅ 환경변수에서 순장 이름을 가져오거나 기본값 사용
CELL_GROUP_LEADERS = {
    "은혜 다락방": os.getenv(
        "LEADERS_EUNHYE", "은혜1,은혜2,은혜3,은혜4,은혜5,은혜6,은혜7,은혜8"
    ).split(","),
    "하품 다락방": os.getenv(
        "LEADERS_HAPOOM", "하품1,하품2,하품3,하품4,하품5,하품6,하품7,하품8"
    ).split(","),
    "오지 다락방": os.getenv(
        "LEADERS_OJI", "오지1,오지2,오지3,오지4,오지5,오지6,오지7,오지8"
    ).split(","),
    "소금 다락방": os.getenv(
        "LEADERS_SOGEUM", "소금1,소금2,소금3,소금4,소금5,소금6,소금7,소금8"
    ).split(","),
    "새가족": os.getenv(
        "LEADERS_NEW", "새가족1,새가족2,새가족3,새가족4,새가족5,새가족6,새가족7,새가족8"
    ).split(","),
}

# ✅ 관리자 비밀번호 (환경변수에서 가져오거나 기본값 사용)
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD")


@app.get("/")
def form_page(request: Request, success: str = Query(default=None)):
    return templates.TemplateResponse(
        "form.html",
        {
            "request": request,
            "cell_groups": list(CELL_GROUP_LEADERS.keys()),
            "cell_group_leaders": CELL_GROUP_LEADERS,
            "success": success,
            "app_version": get_version(),  # 버전 정보 추가
        },
    )


@app.get("/login")
def login_page(request: Request, error: str = None):
    return templates.TemplateResponse(
        "login.html",
        {
            "request": request,
            "error": error,
            "app_version": get_version(),  # 버전 정보 추가
        },
    )


@app.post("/login")
def login(request: Request, password: str = Form(...)):
    if password == ADMIN_PASSWORD:
        # 로그인 성공 시 세션 설정 (간단한 쿠키 기반)
        response = RedirectResponse(url="/admin", status_code=303)
        response.set_cookie(
            key="admin_authenticated",
            value="true",
            max_age=1800,  # 30분으로 단축
            httponly=True,
            secure=os.getenv("ENVIRONMENT") == "production",  # 배포환경에서는 HTTPS 필요
            samesite="lax",
        )
        return response
    else:
        return templates.TemplateResponse(
            "login.html",
            {
                "request": request,
                "error": "비밀번호가 올바르지 않습니다.",
            },
        )


@app.get("/api/leaders/{cell_group}")
def get_leaders_by_cell_group(cell_group: str):
    return {"leaders": CELL_GROUP_LEADERS.get(cell_group, [])}


@app.get("/api/version")
def get_app_version():
    """앱 버전 정보를 반환합니다."""
    return get_build_info()


@app.get("/api/health")
def health_check():
    """UptimeRobot용 헬스체크 엔드포인트 - Cold Start 방지"""
    try:
        # 간단한 DB 연결 테스트
        db = SessionLocal()
        db.execute("SELECT 1")
        db.close()
        
        return {
            "status": "healthy",
            "timestamp": datetime.now(KST).isoformat(),
            "version": get_version(),
            "database": "connected"
        }
    except Exception as e:
        return {
            "status": "unhealthy",
            "timestamp": datetime.now(KST).isoformat(),
            "version": get_version(),
            "database": "disconnected",
            "error": str(e)
        }


@app.post("/submit")
def submit_prayer(
    request: Request,
    name: str = Form(...),
    leader: str = Form(...),
    cell_group: str = Form(...),
    content: str = Form(...),
    is_private: bool = Form(default=False),
):
    try:
        db = SessionLocal()
        new_prayer = Prayer(
            name=name,
            leader=leader,
            cell_group=cell_group,
            content=content,
            created_at=datetime.now(KST),  # ✅ 한국 시간
            is_private=is_private,
        )
        db.add(new_prayer)
        db.commit()
        db.close()
        
        # 📦 캐시 무효화 - 새 기도제목이 등록되면 캐시 삭제
        view_cache.clear()
        view_cache.update({
            "data": None,
            "timestamp": 0,
            "ttl": 30
        })
        
        return RedirectResponse("/?success=true", status_code=303)
    except Exception as e:
        if 'db' in locals():
            db.rollback()
            db.close()
        # 에러 발생 시 에러 페이지로 리다이렉트하는 대신 간단히 재시도 요청
        return RedirectResponse("/?error=true", status_code=303)


def get_week_range(week_offset=0):
    today = datetime.now(KST)  # ✅ 한국 시간
    days_since_sunday = today.weekday() + 1
    if days_since_sunday == 7:
        days_since_sunday = 0
    sunday = today - timedelta(days=days_since_sunday)
    sunday = sunday.replace(hour=0, minute=0, second=0, microsecond=0)
    sunday += timedelta(weeks=week_offset)
    saturday = sunday + timedelta(days=6, hours=23, minutes=59, seconds=59)
    
    # PostgreSQL과의 호환성을 위해 UTC로 변환
    if os.getenv("DATABASE_URL", "").startswith("postgresql"):
        sunday = sunday.astimezone(pytz.UTC)
        saturday = saturday.astimezone(pytz.UTC)
    
    return sunday, saturday


def get_week_label(week_offset=0):
    sunday, saturday = get_week_range(week_offset)
    if sunday.month == saturday.month:
        return f"{sunday.month}월 {sunday.day}일~{saturday.day}일"
    else:
        return f"{sunday.month}월 {sunday.day}일~{saturday.month}월 {saturday.day}일"


def get_cached_prayers_data(week_offset=0):
    """캐싱된 기도제목 데이터를 반환합니다 (30초 캐싱)"""
    current_time = time.time()
    cache_key = f"prayers_{week_offset}"
    
    # 캐시 확인
    if (view_cache.get(cache_key) and 
        current_time - view_cache.get(f"{cache_key}_time", 0) < view_cache["ttl"]):
        return view_cache[cache_key]
    
    # 캐시 만료 또는 없음 - 새로 조회
    db = SessionLocal()
    week_start, week_end = get_week_range(week_offset)
    prayers = (
        db.query(Prayer)
        .filter(Prayer.created_at >= week_start, Prayer.created_at <= week_end)
        .filter(Prayer.is_private.is_(False))  # PostgreSQL boolean 최적화
        .order_by(Prayer.created_at.desc())
        .all()
    )
    db.close()

    # 시간대 변환 - 각 기도제목의 시간을 KST로 변환
    for prayer in prayers:
        if prayer.created_at.tzinfo is None:
            # naive datetime인 경우 UTC로 가정하고 KST로 변환
            prayer.created_at = pytz.UTC.localize(prayer.created_at).astimezone(KST)
        elif prayer.created_at.tzinfo != KST:
            # 이미 timezone이 있는 경우 KST로 변환
            prayer.created_at = prayer.created_at.astimezone(KST)
    
    # 캐시에 저장
    view_cache[cache_key] = prayers
    view_cache[f"{cache_key}_time"] = current_time
    
    return prayers


@app.get("/view")
def view_prayers(
    request: Request,
    leader: str = Query(default=None),
    cell_group: str = Query(default=None),
    week_offset: int = Query(default=0),
):
    # 캐싱된 데이터 사용 (필터링이 없는 경우만)
    if not leader and not cell_group:
        prayers = get_cached_prayers_data(week_offset)
    else:
        # 필터링이 있는 경우 실시간 조회
        db = SessionLocal()
        week_start, week_end = get_week_range(week_offset)
        prayers = (
            db.query(Prayer)
            .filter(Prayer.created_at >= week_start, Prayer.created_at <= week_end)
            .filter(Prayer.is_private.is_(False))
            .order_by(Prayer.created_at.desc())
            .all()
        )
        db.close()

        # 시간대 변환
        for prayer in prayers:
            if prayer.created_at.tzinfo is None:
                prayer.created_at = pytz.UTC.localize(prayer.created_at).astimezone(KST)
            elif prayer.created_at.tzinfo != KST:
                prayer.created_at = prayer.created_at.astimezone(KST)

    unique_leaders = sorted(set(str(p.leader) for p in prayers if p.leader))
    unique_cell_groups = sorted(set(str(p.cell_group) for p in prayers if p.cell_group))

    if leader:
        prayers = [p for p in prayers if str(p.leader) == leader]
    if cell_group:
        prayers = [p for p in prayers if str(p.cell_group) == cell_group]

    grouped = {}
    for p in prayers:
        cg = p.cell_group or "미지정"
        ld = p.leader or "미지정"
        if cg not in grouped:
            grouped[cg] = {}
        if ld not in grouped[cg]:
            grouped[cg][ld] = []
        grouped[cg][ld].append(p)

    week_options = []
    for i in range(-4, 2):
        week_options.append(
            {"offset": i, "label": get_week_label(i), "selected": i == week_offset}
        )

    return templates.TemplateResponse(
        "view.html",
        {
            "request": request,
            "prayers": prayers,
            "leaders": unique_leaders,
            "cell_groups": unique_cell_groups,
            "selected_leader": leader,
            "selected_cell_group": cell_group,
            "grouped": grouped,
            "week_options": week_options,
            "week_offset": week_offset,
            "current_week_label": get_week_label(week_offset),
            "cell_group_leaders": CELL_GROUP_LEADERS,
            "app_version": get_version(),  # 버전 정보 추가
        },
    )


@app.get("/admin")
def view_all(
    request: Request,
    leader: str = Query(default=None),
    cell_group: str = Query(default=None),
    week_offset: int = Query(default=0),
):
    # 인증 체크
    if not request.cookies.get("admin_authenticated"):
        return RedirectResponse(url="/login", status_code=303)
    db = SessionLocal()
    week_start, week_end = get_week_range(week_offset)
    prayers = (
        db.query(Prayer)
        .filter(Prayer.created_at >= week_start, Prayer.created_at <= week_end)
        .order_by(Prayer.created_at.desc())
        .all()
    )
    db.close()

    # 시간대 변환 - 각 기도제목의 시간을 KST로 변환
    for prayer in prayers:
        if prayer.created_at.tzinfo is None:
            # naive datetime인 경우 UTC로 가정하고 KST로 변환
            prayer.created_at = pytz.UTC.localize(prayer.created_at).astimezone(KST)
        elif prayer.created_at.tzinfo != KST:
            # 이미 timezone이 있는 경우 KST로 변환
            prayer.created_at = prayer.created_at.astimezone(KST)

    unique_leaders = sorted(set(str(p.leader) for p in prayers if p.leader))
    unique_cell_groups = sorted(set(str(p.cell_group) for p in prayers if p.cell_group))

    if leader:
        prayers = [p for p in prayers if str(p.leader) == leader]
    if cell_group:
        prayers = [p for p in prayers if str(p.cell_group) == cell_group]

    grouped = {}
    for p in prayers:
        cg = p.cell_group or "미지정"
        ld = p.leader or "미지정"
        if cg not in grouped:
            grouped[cg] = {}
        if ld not in grouped[cg]:
            grouped[cg][ld] = []
        grouped[cg][ld].append(p)

    week_options = []
    for i in range(-4, 2):
        week_options.append(
            {"offset": i, "label": get_week_label(i), "selected": i == week_offset}
        )

    return templates.TemplateResponse(
        "admin.html",
        {
            "request": request,
            "prayers": prayers,
            "leaders": unique_leaders,
            "cell_groups": unique_cell_groups,
            "selected_leader": leader,
            "selected_cell_group": cell_group,
            "grouped": grouped,
            "week_options": week_options,
            "current_week_label": get_week_label(week_offset),
            "cell_group_leaders": CELL_GROUP_LEADERS,
            "app_version": get_version(),  # 버전 정보 추가
        },
    )


@app.get("/export-excel-view")
def export_excel_view(
    request: Request,
    leader: str = Query(default=None),
    cell_group: str = Query(default=None),
    week_offset: int = Query(default=0),
):
    db = SessionLocal()
    week_start, week_end = get_week_range(week_offset)
    prayers = (
        db.query(Prayer)
        .filter(Prayer.created_at >= week_start, Prayer.created_at <= week_end)
        .filter(Prayer.is_private.is_(False))  # PostgreSQL boolean 최적화
        .order_by(Prayer.created_at.desc())
        .all()
    )
    db.close()

    if leader:
        prayers = [p for p in prayers if str(p.leader) == leader]
    if cell_group:
        prayers = [p for p in prayers if str(p.cell_group) == cell_group]

    grouped = {}
    for p in prayers:
        cg = p.cell_group or "미지정"
        ld = p.leader or "미지정"
        if cg not in grouped:
            grouped[cg] = {}
        if ld not in grouped[cg]:
            grouped[cg][ld] = []
        grouped[cg][ld].append(p)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "기도제목 목록"

    cell_group_font = Font(size=16, bold=True)
    leader_font = Font(size=14, bold=True)
    member_font = Font(size=12, bold=True)
    content_font = Font(size=11)

    week_label = get_week_label(week_offset)
    ws.cell(row=1, column=1, value=f"기간: {week_label}")
    ws.cell(row=1, column=1).font = Font(size=14, bold=True)

    row = 3
    for cell_group, leaders in grouped.items():
        ws.cell(row=row, column=1, value=cell_group)
        ws.cell(row=row, column=1).font = cell_group_font
        row += 1

        for leader, prayers in leaders.items():
            ws.cell(row=row, column=2, value=f"{leader} 순장")
            ws.cell(row=row, column=2).font = leader_font
            row += 1

            for p in prayers:
                ws.cell(row=row, column=3, value=p.name)
                ws.cell(row=row, column=3).font = member_font
                ws.cell(row=row, column=4, value=p.content)
                ws.cell(row=row, column=4).font = content_font
                row += 1

    os.makedirs("./static", exist_ok=True)
    filename = (
        f"기도제목목록_{week_label}_{datetime.now(KST).strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    filepath = f"./static/{filename}"
    wb.save(filepath)

    return FileResponse(
        path=filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/export-excel")
def export_excel(
    request: Request,
    leader: str = Query(default=None),
    cell_group: str = Query(default=None),
    week_offset: int = Query(default=0),
):
    # 인증 체크
    if not request.cookies.get("admin_authenticated"):
        return RedirectResponse(url="/login", status_code=303)
    db = SessionLocal()
    week_start, week_end = get_week_range(week_offset)
    prayers = (
        db.query(Prayer)
        .filter(Prayer.created_at >= week_start, Prayer.created_at <= week_end)
        .order_by(Prayer.created_at.desc())
        .all()
    )
    db.close()

    if leader:
        prayers = [p for p in prayers if str(p.leader) == leader]
    if cell_group:
        prayers = [p for p in prayers if str(p.cell_group) == cell_group]

    grouped = {}
    for p in prayers:
        cg = p.cell_group or "미지정"
        ld = p.leader or "미지정"
        if cg not in grouped:
            grouped[cg] = {}
        if ld not in grouped[cg]:
            grouped[cg][ld] = []
        grouped[cg][ld].append(p)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "기도제목 목록"

    cell_group_font = Font(size=16, bold=True)
    leader_font = Font(size=14, bold=True)
    member_font = Font(size=12, bold=True)
    content_font = Font(size=11)

    week_label = get_week_label(week_offset)
    ws.cell(row=1, column=1, value=f"기간: {week_label}")
    ws.cell(row=1, column=1).font = Font(size=14, bold=True)

    row = 3
    for cell_group, leaders in grouped.items():
        ws.cell(row=row, column=1, value=cell_group)
        ws.cell(row=row, column=1).font = cell_group_font
        row += 1

        for leader, prayers in leaders.items():
            ws.cell(row=row, column=2, value=f"{leader} 순장")
            ws.cell(row=row, column=2).font = leader_font
            row += 1

            for p in prayers:
                ws.cell(row=row, column=3, value=p.name)
                ws.cell(row=row, column=3).font = member_font
                ws.cell(row=row, column=4, value=p.content)
                ws.cell(row=row, column=4).font = content_font
                row += 1

    os.makedirs("./static", exist_ok=True)
    filename = (
        f"기도제목목록_{week_label}_{datetime.now(KST).strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    filepath = f"./static/{filename}"
    wb.save(filepath)

    return FileResponse(
        path=filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.delete("/prayer/{prayer_id}")
def delete_prayer(request: Request, prayer_id: int):
    # 인증 체크
    if not request.cookies.get("admin_authenticated"):
        return {"error": "인증이 필요합니다."}
    db = SessionLocal()
    prayer = db.query(Prayer).filter(Prayer.id == prayer_id).first()
    if not prayer:
        db.close()
        return {"error": "기도제목을 찾을 수 없습니다."}
    db.delete(prayer)
    db.commit()
    db.close()
    return {"message": "기도제목이 삭제되었습니다."}


@app.put("/prayer/{prayer_id}")
def update_prayer(
    request: Request,
    prayer_id: int,
    name: str = Form(...),
    leader: str = Form(...),
    cell_group: str = Form(...),
    content: str = Form(...),
    is_private: bool = Form(default=False),
):
    # 인증 체크
    if not request.cookies.get("admin_authenticated"):
        return {"error": "인증이 필요합니다."}
    db = SessionLocal()
    prayer = db.query(Prayer).filter(Prayer.id == prayer_id).first()
    if not prayer:
        db.close()
        return {"error": "기도제목을 찾을 수 없습니다."}

    prayer.name = name
    prayer.leader = leader
    prayer.cell_group = cell_group
    prayer.content = content
    prayer.is_private = is_private
    db.commit()
    db.close()

    return {"message": "기도제목이 수정되었습니다."}
